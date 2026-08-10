"""Financial documents: statements (7 PDFs) + three workbooks.

These get written as plain tables, NOT through the observation/corroboration
model. That is deliberate: unlike the works spine, no independent second
document restates a ledger line or a trial-balance row, so there is nothing to
corroborate against and the observation machinery would add ceremony without
evidence. Where a cross-check DOES exist we assert it inline instead.

TRAP 1 -- UNITS: financial statements declare "All amounts in Lakhs" once, in
a header line, and never repeat it. Read the table without that line and every
figure is out by 100,000. Everything here is converted to rupees.

TRAP 2 -- DO NOT DERIVE: the P&L labels a row "Profit Before Tax (A - B)" but
the stated value never equals Total Revenue minus Total Expenses in any of the
seven years, and the sign differs in most. Read what is stated; never compute
a value the document already gives.
"""
import glob, json, os, re, sys
from pathlib import Path

import openpyxl
from common import text_of, label_map, pick

ROOT = Path(__file__).resolve().parents[1]
DATA = ROOT / "data"

_UNIT_WORDS = {"lakh": 10**5, "lakhs": 10**5, "crore": 10**7, "crores": 10**7,
               "thousand": 10**3, "million": 10**6}

def statement_unit(text):
    """The header declaration: 'INR (All amounts in Lakhs unless stated)'."""
    if m := re.search(r"All amounts in (\w+)", text, re.I):
        return _UNIT_WORDS.get(m.group(1).lower(), 1)
    return 1

def num_after(lines, labels, mult=1):
    """Find a label and take the first NUMERIC line after it.

    Needed because the 2021-2023 era wraps labels across lines: "Total Revenue
    from Operations" then "(A)" then the figure. A plain label->next-line
    lookup returns "(A)". Scanning forward for the first number handles both
    eras with one rule.

    These statements also carry a PREVIOUS YEAR column. The first number after
    the label is the current year, which is what we want.
    """
    for i, line in enumerate(lines):
        norm = re.sub(r"\s+", " ", line).strip()
        if not any(norm.startswith(lab) for lab in labels):
            continue
        for nxt in lines[i + 1:i + 4]:
            v = num(nxt, mult)
            if v is not None:
                return v
    return None

def num(s, mult=1):
    """'(1,234.00)' -> -123400000 with mult=1e5. Brackets mean negative."""
    if s is None:
        return None
    s = str(s).strip()
    neg = s.startswith("(") and s.endswith(")")
    s = s.strip("()").replace(",", "").replace("\u20b9", "").strip()
    if not re.fullmatch(r"-?\d*\.?\d+", s):
        return None
    v = float(s) * mult
    return round(-v if neg else v)

# ------------------------------------------------------------- statements

FS_FIELDS = {
    "revenue_contract":  ["Contract Revenue (EPC)", "Revenue from Contracts"],
    "revenue_total":     ["Total Revenue from Operations", "Total Revenue",
                          "Total Income"],
    "expenses_total":    ["Total Expenses", "Total Expenditure"],
    "profit_before_tax": ["Profit Before Tax", "Profit / (Loss) Before Tax"],
    "tax":               ["Tax Expense (current + deferred)", "Tax Expense",
                          "Provision for Tax"],
    "profit_after_tax":  ["Profit After Tax", "Profit / (Loss) After Tax"],
    # balance-sheet rows are prefixed with their section, e.g.
    # "Current Assets - Trade Receivables", so the bare label never matches
    "paid_up_capital":   ["Shareholders' Funds \u2014 Paid-up Capital",
                          "Paid-up Share Capital", "Share Capital"],
    "reserves":          ["Reserves & Surplus", "Reserves and Surplus"],
    "borrowings":        ["Non-Current Liabilities \u2014 Long-term Borrowings",
                          "Borrowings", "Total Borrowings"],
    "payables":          ["Current Liabilities \u2014 Trade Payables"],
    "inventories":       ["Current Assets \u2014 Inventories"],
    "receivables":       ["Current Assets \u2014 Trade Receivables", "Trade Receivables"],
    "cash":              ["Current Assets \u2014 Cash & Bank Balances",
                          "Cash & Cash Equivalents"],
}

def statements(root):
    rows = []
    for path in sorted(glob.glob(os.path.join(root, "documents/financial_statement/*.pdf"))):
        text = text_of(path)
        mult = statement_unit(text)
        lines = [l.strip() for l in text.split("\n") if l.strip()]
        doc = os.path.basename(path)[:-4]
        # the document's own FY, not the one in the filename -- DOC-FS-2022
        # reports FY2022-23, so the filename year alone is misleading
        fy = (m.group(1) if (m := re.search(r"\(FY(\d{4}-\d{2})\)", text))
              else (m2.group(1) if (m2 := re.search(r"year ended 31st March (\d{4})", text))
                    else None))
        row = {"doc": doc, "fiscal_year": fy, "unit_multiplier": mult}
        for field, labels in FS_FIELDS.items():
            row[field] = num_after(lines, labels, mult)
        # stated internal consistency: PBT - tax = PAT (this one DOES hold)
        if all(row.get(k) is not None for k in ("profit_before_tax", "tax", "profit_after_tax")):
            row["pat_check"] = abs(row["profit_before_tax"] - row["tax"]
                                   - row["profit_after_tax"]) <= mult
        rows.append(row)
    return rows

# ------------------------------------------------------------- workbooks

def sheet_rows(path, sheet):
    """Header row -> list of dicts. Workbook headers are clean, no coordinates
    needed (unlike the PDF tables)."""
    ws = openpyxl.load_workbook(path, data_only=True)[sheet]
    it = ws.iter_rows(values_only=True)
    header = [str(c).strip() if c is not None else "" for c in next(it)]
    out = []
    for r in it:
        if all(c is None for c in r):
            continue
        out.append({h: v for h, v in zip(header, r) if h})
    return out

def trial_balance(root):
    path = os.path.join(root, "documents/workbooks/Trial_Balance_by_Year.xlsx")
    wb = openpyxl.load_workbook(path, data_only=True)
    rows = []
    for sheet in wb.sheetnames:
        if not sheet.startswith("TB "):
            continue
        for r in sheet_rows(path, sheet):
            rows.append({"fiscal_year": sheet.replace("TB ", ""),
                         "account": r.get("Account"),
                         "debit": num(r.get("Debit (INR)")),
                         "credit": num(r.get("Credit (INR)")),
                         "balance": num(r.get("Balance (INR)"))})
    return rows

def receivables(root):
    path = os.path.join(root, "documents/workbooks/Receivables_Ageing.xlsx")
    rows = []
    for r in sheet_rows(path, "AR Ageing"):
        rows.append({"invoice_no": r.get("Invoice No"),
                     "client": r.get("Client"),
                     "invoice_date": str(r.get("Invoice Date"))[:10] if r.get("Invoice Date") else None,
                     "invoiced": num(r.get("Invoiced (INR)")),
                     "received": num(r.get("Received (INR)")),
                     "outstanding": num(r.get("Outstanding (INR)")),
                     "status": r.get("Status")})
    return rows

def assets(root):
    path = os.path.join(root, "documents/workbooks/Plant_and_Machinery_Register.xlsx")
    rows = []
    for r in sheet_rows(path, "Plant Register"):
        rows.append({"asset_id": r.get("Asset ID"), "type": r.get("Type"),
                     "make": r.get("Make"), "acquired": r.get("Acquired"),
                     "cost": num(r.get("Cost (INR)")), "condition": r.get("Condition"),
                     "location": r.get("Location"), "ownership": r.get("Ownership")})
    return rows


if __name__ == "__main__":
    root = sys.argv[1] if len(sys.argv) > 1 else "../BITS-Hackathon-Dataset"
    DATA.mkdir(exist_ok=True)

    fs, tb, ar, ass = statements(root), trial_balance(root), receivables(root), assets(root)

    print(f"financial statements {len(fs)}")
    for r in fs:
        miss = [k for k in FS_FIELDS if r.get(k) is None]
        print(f"   {r['doc']:14s} unit x{r['unit_multiplier']:<8} PAT check={r.get('pat_check')}"
              f"   missing: {miss if miss else 'none'}")

    print(f"\ntrial balance rows  {len(tb)}   years {sorted({r['fiscal_year'] for r in tb})}")
    print(f"receivables         {len(ar)}   clients {len({r['client'] for r in ar})}")
    print(f"   invoiced total INR {sum(r['invoiced'] or 0 for r in ar)/1e7:,.1f} Cr"
          f"   received INR {sum(r['received'] or 0 for r in ar)/1e7:,.1f} Cr")
    print(f"assets              {len(ass)}  owned {sum(1 for a in ass if a['ownership']=='owned')}"
          f"   cost INR {sum(a['cost'] or 0 for a in ass)/1e7:,.1f} Cr")

    # the compliance matrices claim 210 owned assets and 486 personnel --
    # a checkable cross-document assertion
    # the compliance matrices claim "210 owned assets". The register holds 211
    # rows: 154 owned, 56 leased, 1 blank. So the claim matches the TOTAL, not
    # the owned subset -- the matrix wording is loose. Do not use it as a
    # checksum for ownership; count from the register.
    import collections as _c
    print(f"\n   asset ownership: {dict(_c.Counter(a['ownership'] for a in ass))}"
          f"   (compliance matrices claim '210 owned' = total, not the owned subset)")

    for name, rows in [("financials", fs), ("trial_balance", tb),
                       ("receivables", ar), ("assets", ass)]:
        json.dump(rows, open(DATA / f"{name}.json", "w"), indent=1)
    print("\nwrote financials.json, trial_balance.json, receivables.json, assets.json")
